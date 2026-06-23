import cv2
import numpy as np
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import fitz  # PyMuPDF
import re
from tkinter import filedialog, messagebox
import time
import sys
# =========================
# CONFIG
# =========================
excel_path = r"D:\AI_CODING\PYTHON\project1\Quote_Generator\Quote_Book.xlsx"

SCALE = 0.3
CM_PER_PIXEL = 2.54 / (300 * SCALE)

import math   # (you already have this, so ignore if duplicate)

def round_cm_custom(v):
    base = math.floor(v)
    decimal = v - base

    if decimal <= 0.25:
        return float(base)

    elif decimal <= 0.51:
        return base + 0.5

    else:
        return float(base + 1)

def get_cm_offset():
    """Convert entry cm → pixel offset"""
    try:
        h_cm = float(h_entry.get() or 0)
    except:
        h_cm = 0

    try:
        w_cm = float(w_entry.get() or 0)
    except:
        w_cm = 0

    h_px = int(h_cm / CM_PER_PIXEL)
    w_px = int(w_cm / CM_PER_PIXEL)

    return w_px, h_px

# =========================
# GLOBAL STATE
# =========================
global selected_box_index, selected_mark_index, selected_mark_source
img_cv = None
img_display = None
canvas_img = None

history_stack = []
redo_stack = []

start_x, start_y = 0, 0
dragging = False

selected_ids = set()
selected_mark_ids = set()

pdf_doc = None
page_index = 0
total_pages = 0

current_file_path = None
page_names = []
page_data = {}  # key: page_index, value: history_stack

global zoom_scale
zoom_scale = 1.0

preview_rect = None

selected_box_index = None
drag_offset_x = 0
drag_offset_y = 0
zoom_scale = 1.0
dragging = False
panning = False

HANDLE_SIZE = 6
resize_mode = False
resize_corner = None

selected_box_index = None
resize_mode = False
resize_corner = None
move_mode = False
edit_enabled = False
global_marks = []   # shared across ALL pages
mounting_marks = []
page_marks = {}
mark_mode = False
selected_mark_index = None
saved_flag = False
done_click_time = 0
selected_mark_is_page = False
selected_mark_source = None
CONFIG_FILE = "customer_link.txt"
last_redraw_time = 0
img_rgb_cache = None

# ===== NEW MODE STATE =====
mode = "cut"
select_mode = False

# =========================
# MODE SWITCH
# =========================
def set_mode(m):
    global mode
    mode = m
    print("Mode:", mode)

# =========================
# ZOOM FUNCTIONS (NEW)
# =========================
def zoom_in():
    global zoom_scale
    zoom_scale += 0.1
    redraw()

def zoom_out():
    global zoom_scale
    zoom_scale = max(0.2, zoom_scale - 0.1)
    redraw()

def reset_zoom():
    global zoom_scale
    zoom_scale = 1.0
    redraw()

def find_box(x, y):
    """Return index of box under cursor"""
    for i in range(len(history_stack)-1, -1, -1):
        (x1, y1, x2, y2), _ = history_stack[i]
        if x1 <= x <= x2 and y1 <= y <= y2:
            return i
    return None


def get_handle(x, y, box):
    """Check if clicking near a corner (resize handle)"""
    x1, y1, x2, y2 = box

    handles = {
        "tl": (x1, y1),
        "tr": (x2, y1),
        "bl": (x1, y2),
        "br": (x2, y2)
    }

    for name, (hx, hy) in handles.items():
        if abs(x - hx) < HANDLE_SIZE and abs(y - hy) < HANDLE_SIZE:
            return name
    return None

    
# =========================
# MOUSE WHEEL ZOOM (ADD HERE)
# =========================
def on_mouse_wheel(event):
    global zoom_scale, dragging

    # 🚫 prevent zoom while dragging selection
    if dragging:
        return

    if event.delta > 0:
        zoom_scale += 0.1
    else:
        zoom_scale = max(0.2, zoom_scale - 0.1)

    redraw()

# =========================
# PAN FUNCTIONS (ADD HERE)
# =========================
def start_pan(event):
    global panning
    panning = True
    canvas.scan_mark(event.x, event.y)

def do_pan(event):
    if panning:
        canvas.scan_dragto(event.x, event.y, gain=1)

def stop_pan(event):
    global panning
    panning = False

# =========================
# AUTO CUT FUNCTION
# =========================
def auto_cut():
    global history_stack, selected_ids, selected_mark_ids, global_marks, page_marks

    if img_cv is None:
        return


    objects = get_objects(img_cv)

    px = int(1 / CM_PER_PIXEL)

    # =========================
    # ADD MARKS AS OBJECTS (FIXED)
    # =========================
    fake_id = 100000

    for mx, my in global_marks:

        # 🔥 FIX: DO NOT clear before checking usage
        if (mx, my) in selected_mark_ids:
            continue

        x1 = mx - px // 2
        y1 = my - px // 2
        x2 = mx + px // 2
        y2 = my + px // 2

        # 🔥 FIX: unique ID per mark (important)
        objects.append((fake_id, x1, y1, x2, y2))

        fake_id += 1

    selected = []
    obj_ids = set()
    selected_marks = []

    for obj_id, x1, y1, x2, y2 in objects:
        if obj_id in selected_ids:
            continue

        selected.append((x1, y1, x2, y2))
        obj_ids.add(obj_id)

    if not selected:
        return

    xs, ys = [], []

    for x1, y1, x2, y2 in selected:
        xs += [x1, x2]
        ys += [y1, y2]

    bx1, by1, bx2, by2 = min(xs), min(ys), max(xs), max(ys)

    # =========================
    # MOUNTING MARK MODE
    # =========================
    added_marks = []

    if mode_var.get() == "Mounting Mark":

        gap_px = int(1.5 / CM_PER_PIXEL)

        cy = (by1 + by2) // 2

        left_mark = (int(bx1 - gap_px), int(cy))
        right_mark = (int(bx2 + gap_px), int(cy))

        page_marks.setdefault(page_index, [])

        if left_mark not in page_marks[page_index]:
            page_marks[page_index].append(left_mark)
            added_marks.append(left_mark)

        if right_mark not in page_marks[page_index]:
            page_marks[page_index].append(right_mark)
            added_marks.append(right_mark)

        bx1 = min(bx1, left_mark[0] - px // 2)
        bx2 = max(bx2, right_mark[0] + px // 2)
        by1 = min(by1, cy - px // 2)
        by2 = max(by2, cy + px // 2)

    # =========================
    # OFFSET
    # =========================
    w_off, h_off = get_cm_offset()

    bx1 -= w_off
    bx2 += w_off
    by1 -= h_off
    by2 += h_off

    # =========================
    # FINAL STATE UPDATE
    # =========================
    selected_ids.update(obj_ids)

    # 🔥 FIX: mark locking happens AFTER selection
    for mx, my in global_marks:
        for obj_id, x1, y1, x2, y2 in objects:
            if x1 <= mx <= x2 and y1 <= my <= y2:
                selected_mark_ids.add((mx, my))

    history_stack.append((
        (bx1, by1, bx2, by2),
        obj_ids,
        selected_marks,
        added_marks
    ))

    redraw()

# =========================
# EXTRACT PAGE NAMES
# =========================
def extract_page_names(file_path):
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    try:
        last_part = file_name.split("_")[-1]
        parts = last_part.split("-")
        return [p.strip() for p in parts if p.strip()]
    except:
        return []

def build_clean_filename(full_name, page_name):
    base = os.path.splitext(os.path.basename(full_name))[0]

    # STEP 1: remove everything after last "_"
    base = base.rsplit("_", 1)[0]

    # STEP 2: add page name
    page_name = page_name.strip()

    return f"{base}_{page_name}"

# =========================
# FILE OPEN
# =========================
def open_file():
    global current_file_path, saved_flag

    # =========================
    # PICK FILE ONLY ONCE
    # =========================
    file_path = filedialog.askopenfilename(
        filetypes=[
            ("All Supported", "*.png *.jpg *.jpeg *.pdf"),
            ("Images", "*.png *.jpg *.jpeg"),
            ("PDF", "*.pdf")
        ]
    )

    if not file_path:
        file_name_var.set("")
        return

    # =========================
    # UNSAVED CHECK
    # =========================
    if has_unsaved_changes():
        response = messagebox.askyesno(
            "Unsaved Changes",
            "You have unsaved changes.\nDo you want to open a new file anyway?"
        )

        if not response:
            return

        full_reset()

    # =========================
    # SET FILE
    # =========================
    current_file_path = file_path
    saved_flag = False

    # =========================
    # SHOW FILE NAME
    # =========================
    raw_name = os.path.splitext(os.path.basename(current_file_path))[0]
    base_name = raw_name.rsplit("_", 1)[0]
    base_name_var.set(base_name)

    # =========================
    # LOAD FILE
    # =========================
    if file_path.lower().endswith(".pdf"):
        load_pdf(file_path)
    else:
        load_image(file_path)

def close_file():
    global img_cv, img_display, pdf_doc
    global history_stack, redo_stack, selected_ids
    global current_file_path, page_index, total_pages
    global saved_flag   # ✅ FIXED
    
    file_name_var.set("")
    base_name_var.set("")
    
    # =========================
    # SAVE CHECK (NEW LOGIC)
    # =========================
    if not saved_flag:
        response = messagebox.askyesnocancel(
            "Save File",
            "You have not saved your work.\nDo you want to save before closing?"
        )
    
        if response is None:
            return  # Cancel
    
        if response:  # YES → Save then close
            save_all()
    
    # =========================
    # RESET ALL STATE
    # =========================
    saved_flag = True   # treat next open as clean
    
    img_cv = None
    img_display = None
    pdf_doc = None
    
    history_stack.clear()
    redo_stack.clear()
    selected_ids.clear()
    
    current_file_path = None
    page_index = 0
    total_pages = 0
    
    # 🔥 ADD THIS (IMPORTANT)
    page_data.clear()
    page_marks.clear()
    global_marks.clear()
    selected_mark_ids.clear()
    
    page_label_var.set("Color Name")
    canvas.delete("all")

    img_cv = None
    img_display = None
    pdf_doc = None

    history_stack.clear()
    redo_stack.clear()
    selected_ids.clear()

    current_file_path = None
    page_index = 0
    total_pages = 0

    page_label_var.set("Color Name")
    canvas.delete("all")
    # =========================
    # RESET ALL MODES + BUTTON UI
    # =========================
    reset_all_modes()

    # clear selections
    global selected_box_index
    global selected_mark_index
    global selected_mark_source

    selected_box_index = None
    selected_mark_index = None
    selected_mark_source = None

    # reset mode dropdown
    mode_var.set("No Mark")

    # clear preview
    update_preview_box(None)

    # clear resize entries
    h_entry2.delete(0, tk.END)
    w_entry2.delete(0, tk.END)

    # disable resize panel
    set_framebox2_state(False)

    # reset zoom
    global zoom_scale
    zoom_scale = 1.0

    # clear sidebar color buttons
    for widget in sidebar_btn_frame.winfo_children():
        widget.destroy()
    # =========================
    # CLEAR FRAMEBOX VALUES
    # =========================
    h_entry.delete(0, tk.END)
    w_entry.delete(0, tk.END)

    h_entry2.delete(0, tk.END)
    w_entry2.delete(0, tk.END)

    # =========================
    # CLEAR PLATE + EDITOR
    # =========================
    plate_var.set("")
    editor_var.set("")
    # remove focus from entries/buttons
    root.focus_set()

    print("File closed ✔")
    

# =========================
# LOAD IMAGE
# =========================
def load_image(path):
    global img_cv, img_display, history_stack, redo_stack, selected_ids, pdf_doc

    pdf_doc = None

    img = Image.open(path).convert("RGB")
    img = np.array(img)

    img_cv = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    img_cv = cv2.resize(img_cv, (0, 0), fx=SCALE, fy=SCALE)

    img_display = img_cv.copy()

    history_stack.clear()
    redo_stack.clear()
    selected_ids.clear()

    edit_btn.config(state="normal") 

    page_label_var.set(os.path.basename(path))
    redraw()

# =========================
# PDF HANDLING
# =========================
def load_pdf(path):
    global pdf_doc, page_index, total_pages, page_names

    pdf_doc = fitz.open(path)
    total_pages = len(pdf_doc)
    page_index = 0

    page_names = extract_page_names(path)

    build_sidebar()   # ✅ ADD THIS LINE

    load_pdf_page(page_index)

def build_sidebar():
    for widget in sidebar_btn_frame.winfo_children():
        widget.destroy()

    if not pdf_doc:
        return

    btn_all = tk.Button(
        sidebar_btn_frame,
        text="ALL\nCOLORS",
        bg="#40aac7",
        fg="white",
        command=lambda: go_to_page(0)
    )
    btn_all.pack(fill="x", padx=5, pady=5)

    for i in range(1, total_pages):
        name = page_names[i - 1] if (i - 1) < len(page_names) else f"Sep {i}"

        btn = tk.Button(
            sidebar_btn_frame,
            text=name,
            command=lambda idx=i: go_to_page(idx)
        )
        btn.pack(fill="x", padx=5, pady=2)

        

def go_to_page(index):
    global page_index

    if has_unsaved_changes():
        set_status("⚠ Click DONE before changing page", True)
        return

    save_current_page_state()
    page_index = index
    load_pdf_page(page_index)

def save_current_page_state():
    global page_data

    page_data[page_index] = {
        "history": history_stack.copy(),
        "redo": redo_stack.copy(),
        "selected": selected_ids.copy(),
        "marks": global_marks.copy()   # ✅ ADD THIS
    }
    
def restore_page_state():
    global history_stack, redo_stack, selected_ids, global_marks

    if page_index not in page_data:
        page_data[page_index] = {
            "history": [],
            "redo": [],
            "selected": set(),
            "marks": []
        }

    state = page_data[page_index]
    
    history_stack[:] = state.get("history", [])
    redo_stack[:] = state.get("redo", [])
    selected_ids.clear()
    selected_ids.update(state.get("selected", set()))

def load_pdf_page(index):
    global img_cv, img_display, page_index, selected_mark_ids

    page_index = index
    page = pdf_doc.load_page(page_index)

    pix = page.get_pixmap(dpi=300)
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

    if pix.n == 4:
        img = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)

    img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)

    # =========================
    # SCALE ORIGINAL PAGE
    # =========================
    img = cv2.resize(img, (0, 0), fx=SCALE, fy=SCALE)

    # =========================
    # ADD 20 CM PADDING AROUND PAGE
    # =========================
    pad_px = int(10 / CM_PER_PIXEL)

    img_cv = cv2.copyMakeBorder(
        img,
        pad_px, pad_px, pad_px, pad_px,
        borderType=cv2.BORDER_CONSTANT,
        value=(255, 255, 255)   # white background
    )

    img_display = img_cv.copy()

    # reset mark selection lock when changing page
    selected_mark_ids.clear()

    # restore AFTER image is ready
    restore_page_state()

    # page label
    if page_index == 0:
        current_name = "ALL COLORS"
    else:
        if (page_index - 1) < len(page_names):
            current_name = page_names[page_index - 1]
        else:
            current_name = f"Sep {page_index}"

    set_status(current_name)

    edit_btn.config(state="normal")

    redraw()

def has_unsaved_changes():
    global page_index

    if not history_stack and not global_marks:
        return False

    if page_index not in page_data:
        return True

    state = page_data[page_index]

    saved_history = state.get("history", [])

    if len(history_stack) != len(saved_history):
        return True

    return False
    
# =========================
# page select
# =========================
def next_page():
    global page_index, selected_mark_ids  # ✅ include this

    if pdf_doc and page_index < total_pages - 1:

        if has_unsaved_changes():
            set_status("⚠ Click DONE before changing page", True)
            return

        save_current_page_state()

        # ✅🔥 FIX: reset mark selection lock
        selected_mark_ids.clear()

        page_index += 1
        load_pdf_page(page_index)

def prev_page():
    global page_index

    if pdf_doc and page_index > 0:

        if has_unsaved_changes():
            set_status("⚠ Click DONE before changing page", True)
            return

        save_current_page_state()
        # ✅🔥 FIX: reset mark selection lock
        selected_mark_ids.clear()
        page_index -= 1
        load_pdf_page(page_index)

# =========================
# OBJECT DETECTION
# =========================
def get_objects(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    mask = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        11,
        2
    )
    
    kernel = np.ones((3, 3), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    objects = []

    for i, c in enumerate(contours):
        x, y, w, h = cv2.boundingRect(c)

        # 🔥 ADD THIS (noise filter)
        area = w * h
        if area < 20:     # adjust 10–50 depending on your noise level
            continue

        if w < 2 or h < 2:
            continue

        x2, y2 = x + w, y + h
        objects.append((i, x, y, x2, y2))

    return objects

def process_area(x1, y1, x2, y2):
    global history_stack, redo_stack, selected_ids, selected_mark_ids, global_marks, page_marks, page_index

    added_marks = []

    # 🚫 GLOBAL HARD LOCK
    IGNORE_MARKS = (mode_var.get() == "Mounting Mark")

    # =========================
    # normalize coordinates FIRST
    # =========================
    x1, x2 = min(x1, x2), max(x1, x2)
    y1, y2 = min(y1, y2), max(y1, y2)

    objects = get_objects(img_cv)

    selected = []
    obj_ids = set()

    for obj_id, bx1, by1, bx2, by2 in objects:
        if obj_id in selected_ids:
            continue

        if bx1 >= x1 and by1 >= y1 and bx2 <= x2 and by2 <= y2:
            selected.append((bx1, by1, bx2, by2))
            obj_ids.add(obj_id)

    # =========================
    # INCLUDE MARKS INSIDE BOX
    # =========================
    selected_marks = []

    px = int(1 / CM_PER_PIXEL)

    for i, (mx, my) in enumerate(global_marks):

        if mode_var.get() == "Mounting Mark":
            continue

        # 🔥 FIX: prevent already-used marks (stable logic)
        if (mx, my) in selected_mark_ids:
            continue

        bx1 = mx - px // 2
        bx2 = mx + px // 2
        by1 = my - px // 2
        by2 = my + px // 2

        if not (bx2 < x1 or bx1 > x2 or by2 < y1 or by1 > y2):
            selected_marks.append((mx, my))
            selected_mark_ids.add((mx, my))   # 🔥 FIXED (was index-based before)

    # ❌ nothing selected
    if not selected and not selected_marks:
        return False

    selected_ids.update(obj_ids)

    # =========================
    # compute bounding box base
    # =========================
    xs, ys = [], []

    for bx1, by1, bx2, by2 in selected:
        xs += [bx1, bx2]
        ys += [by1, by2]

    px = int(1 / CM_PER_PIXEL)

    for mx, my in selected_marks:
        xs += [mx - px // 2, mx + px // 2]
        ys += [my - px // 2, my + px // 2]

    # =========================
    # MOUNTING MARK LOGIC
    # =========================
    added_marks = []

    if mode_var.get() == "Mounting Mark":

        gap_px = int(1.5 / CM_PER_PIXEL)

        base_x1, base_y1, base_x2, base_y2 = min(xs), min(ys), max(xs), max(ys)

        cy = (base_y1 + base_y2) // 2

        MIN_W_CM = 12
        min_w_px = int(MIN_W_CM / CM_PER_PIXEL)

        current_w = base_x2 - base_x1

        if current_w < min_w_px:
            diff = min_w_px - current_w
            base_x1 -= diff // 2
            base_x2 += diff // 2

        left_mark = (int(base_x1 - gap_px), int(cy))
        right_mark = (int(base_x2 + gap_px), int(cy))

        page_marks.setdefault(page_index, [])

        if left_mark not in page_marks[page_index]:
            page_marks[page_index].append(left_mark)
            added_marks.append(left_mark)

        if right_mark not in page_marks[page_index]:
            page_marks[page_index].append(right_mark)
            added_marks.append(right_mark)

        xs += [
            left_mark[0] - px // 2,
            left_mark[0] + px // 2,
            right_mark[0] - px // 2,
            right_mark[0] + px // 2,
        ]

        ys += [
            left_mark[1] - px // 2,
            left_mark[1] + px // 2,
            right_mark[1] - px // 2,
            right_mark[1] + px // 2,
        ]

    # =========================
    # FINAL BOX
    # =========================
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)

    # =========================
    # CM OFFSET
    # =========================
    w_off, h_off = get_cm_offset()

    x1 -= w_off
    x2 += w_off
    y1 -= h_off
    y2 += h_off

    history_stack.append((
        (x1, y1, x2, y2),
        obj_ids,
        selected_marks,
        added_marks
    ))

    redo_stack.clear()
    redraw()

    return True

# =========================
# UNDO / REDO
# =========================
def undo():
    global history_stack, redo_stack, selected_ids
    if not history_stack:
        return
    last = history_stack.pop()
    redo_stack.append(last)
    
    # 🔥 HANDLE MARK UNDO
    if last[0] == "mark":
        mark = last[1]
        if mark in global_marks:
            global_marks.remove(mark)
        redraw()
        return
    
    # ===== existing box logic =====
    _, obj_ids, selected_marks, *rest = last
    
    added_marks = rest[0] if rest else []
    selected_ids -= obj_ids
    
    # 🔥 unlock all marks from this undo item
    for mx, my in selected_marks:
        selected_mark_ids.discard((mx, my))
    
    # 🔥 IMPORTANT: auto_cut marks are stored differently
    for mx, my in global_marks:
        inside = False
    
        for item in history_stack:
    
            if item[0] == "mark":
                continue
    
            box = item[0]
            bx1, by1, bx2, by2 = box
    
            if bx1 <= mx <= bx2 and by1 <= my <= by2:
                inside = True
                break
    
        # if no remaining box uses this mark → unlock it
        if not inside:
            selected_mark_ids.discard((mx, my))
    # 🔥 remove mounting marks added by this action
    for mark in added_marks:
    
        # remove from global marks if exists
        if mark in global_marks:
            try:
                global_marks.remove(mark)
            except:
                pass
    
        # 🔥 ALSO REMOVE FROM PAGE MARKS (THIS FIXES YOUR ISSUE)
        if page_index in page_marks:
            if mark in page_marks[page_index]:
                page_marks[page_index].remove(mark)  
    
    redraw()

def redo():
    global history_stack, redo_stack, selected_ids
    if not redo_stack:
        return

    item = redo_stack.pop()
    history_stack.append(item)

    # 🔥 HANDLE MARK REDO
    if item[0] == "mark":
        mark = item[1]
        if mark not in global_marks:
            global_marks.append(mark)
        redraw()
        return

    # ===== existing box logic =====
    obj_ids = item[1]
    selected_marks = item[2]
    added_marks = item[3] if len(item) > 3 else []

    selected_ids.update(obj_ids)

    # 🔥 re-add mounting marks
    for mark in added_marks:
        if page_index not in page_marks:
            page_marks[page_index] = []
    
        if mark not in page_marks[page_index]:
            page_marks[page_index].append(mark)

    redraw()
    # ===== existing box logic =====
    _, obj_ids, *rest = item
    
    added_marks = rest[1] if len(rest) > 1 else []
    selected_ids.update(obj_ids)

    # 🔥 re-add mounting marks
    for mark in added_marks:
        if mark not in global_marks:
            global_marks.append(mark)
    
    redraw()

def redraw():
    global last_redraw_time

    now = time.time()
    if now - last_redraw_time < 0.03:   # ~30 FPS limit
        return

    last_redraw_time = now

    global img_display

    if img_cv is None:
        return

    img_display = img_cv.copy()

    # =========================
    # LOCAL SPEED CACHE (IMPORTANT)
    # =========================
    hist = history_stack
    sel_index = selected_box_index
    edit = edit_enabled
    zoom = zoom_scale

    cm_px = CM_PER_PIXEL
    marks_global = global_marks
    marks_page = page_marks.get(page_index, [])

    # =========================
    # DRAW BOXES
    # =========================
    for i, item in enumerate(hist):

        if not item or item[0] == "mark":
            continue

        box = item[0]
        if len(box) != 4:
            continue

        try:
            x1, y1, x2, y2 = box
        except:
            continue

        color = (0, 0, 255)
        thickness = 2

        if sel_index == i:
            if edit:
                color = (0, 255, 255)
            else:
                color = (255, 0, 255)
            thickness = 3

        cv2.rectangle(img_display, (x1, y1), (x2, y2), color, thickness)

    # =========================
    # DRAW HANDLES (EDIT MODE ONLY)
    # =========================
    if edit:
        handle_size = int(10 / zoom) + 4

        for item in hist:
            if not item or item[0] == "mark":
                continue

            box = item[0]
            if len(box) != 4:
                continue

            x1, y1, x2, y2 = box

            cx = (x1 + x2) // 2
            cy = (y1 + y2) // 2

            cv2.circle(img_display, (cx, y1), handle_size, (250, 67, 0), -1)
            cv2.circle(img_display, (cx, y2), handle_size, (250, 67, 0), -1)
            cv2.circle(img_display, (x1, cy), handle_size, (250, 67, 0), -1)
            cv2.circle(img_display, (x2, cy), handle_size, (250, 67, 0), -1)

    # =========================
    # DRAW GLOBAL MARKS
    # =========================
    px = int(1 / cm_px)
    thickness = 2

    sel_mark_source = globals().get("selected_mark_source")
    sel_mark_index = globals().get("selected_mark_index")

    for i, (x, y) in enumerate(marks_global):

        if select_mode and sel_mark_source == "global" and sel_mark_index == i:
            color = (255, 0, 255)
            t = 3
        else:
            color = (0, 0, 255)
            t = thickness

        cv2.line(img_display, (x - px // 2, y), (x + px // 2, y), color, t)
        cv2.line(img_display, (x, y - px // 2), (x, y + px // 2), color, t)

    # =========================
    # ACTIVE MARKS CACHE
    # =========================
    active_marks = set()

    if sel_index is not None and sel_index < len(hist):
        item = hist[sel_index]
        if len(item) > 3:
            active_marks = set(item[3])

    # =========================
    # DRAW PAGE MARKS
    # =========================
    for x, y in marks_page:

        color = (255, 0, 0)

        if (x, y) in active_marks:
            color = (255, 255, 0)

        cv2.line(img_display, (x - px // 2, y), (x + px // 2, y), color, 2)
        cv2.line(img_display, (x, y - px // 2), (x, y + px // 2), color, 2)

    # =========================
    # FINAL RENDER
    # =========================
    draw_image()

# =========================
# DRAW IMAGE
# =========================
def draw_image():
    global canvas_img

    if img_display is None:
        return

    canvas.delete("all")

    img_rgb = cv2.cvtColor(img_display, cv2.COLOR_BGR2RGB)
    img_pil = Image.fromarray(img_rgb)

    # FIT IMAGE TO CANVAS (base scale)
    canvas_w = canvas.winfo_width()
    canvas_h = canvas.winfo_height()

    img_h, img_w = img_cv.shape[:2]

    fit_scale = min(canvas_w / img_w, canvas_h / img_h)

    # APPLY USER ZOOM ON TOP OF FIT
    final_scale = fit_scale * zoom_scale

    new_w = int(img_w * final_scale)
    new_h = int(img_h * final_scale)

    img_pil = img_pil.resize((new_w, new_h), Image.Resampling.LANCZOS)

    canvas_img = ImageTk.PhotoImage(img_pil)

    # ✅ CENTER IMAGE
    x_center = canvas_w // 2
    y_center = canvas_h // 2

    canvas.create_image(x_center, y_center, anchor="center", image=canvas_img)

    # ✅ FIX SCROLL REGION FOR CENTERED IMAGE
    canvas.config(scrollregion=(
        x_center - new_w // 2,
        y_center - new_h // 2,
        x_center + new_w // 2,
        y_center + new_h // 2
    ))

    
def screen_to_image(x, y):
    # canvas coordinates (after pan)
    cx = canvas.canvasx(x)
    cy = canvas.canvasy(y)

    canvas_w = canvas.winfo_width()
    canvas_h = canvas.winfo_height()

    img_h, img_w = img_cv.shape[:2]

    fit_scale = min(canvas_w / img_w, canvas_h / img_h)
    total_scale = fit_scale * zoom_scale

    # ✅ OFFSET because image is CENTERED
    x_offset = (canvas_w - img_w * total_scale) / 2
    y_offset = (canvas_h - img_h * total_scale) / 2

    # ✅ subtract offset BEFORE scaling back
    img_x = (cx - x_offset) / total_scale
    img_y = (cy - y_offset) / total_scale

    return int(img_x), int(img_y)

def find_box(x, y):
    for i in range(len(history_stack) - 1, -1, -1):

        item = history_stack[i]

        # 🔥 skip mark entries
        if item[0] == "mark":
            continue

        (x1, y1, x2, y2), *_ = item

        if x1 <= x <= x2 and y1 <= y <= y2:
            return i

    return None
    
def is_on_border(x, y, box):
    x1, y1, x2, y2 = box

    tol = max(10, int(20 / zoom_scale))  # clickable thickness

    # inside expanded box
    if x1 - tol <= x <= x2 + tol and y1 - tol <= y <= y2 + tol:
        # but NOT deep inside → only border
        if not (x1 + tol < x < x2 - tol and y1 + tol < y < y2 - tol):
            return True

    return False

def get_resize_direction(x, y, box):
    x1, y1, x2, y2 = box

    tol = max(10, int(20 / zoom_scale))

    corners = {
        "tl": (x1, y1),
        "tr": (x2, y1),
        "bl": (x1, y2),
        "br": (x2, y2)
    }

    # check corners first (highest priority)
    for name, (cx, cy) in corners.items():
        if abs(x - cx) <= tol and abs(y - cy) <= tol:
            return name

    # fallback to edges
    dx_left = abs(x - x1)
    dx_right = abs(x - x2)
    dy_top = abs(y - y1)
    dy_bottom = abs(y - y2)

    min_dist = min(dx_left, dx_right, dy_top, dy_bottom)

    if min_dist == dx_left:
        return "left"
    elif min_dist == dx_right:
        return "right"
    elif min_dist == dy_top:
        return "top"
    else:
        return "bottom"
        
    
# draw handles only in edit mode
if edit_enabled:
    for (box, _) in history_stack:
        x1, y1, x2, y2 = box

        handles = [
            ((x1 + x2) // 2, y1),
            ((x1 + x2) // 2, y2),
            (x1, (y1 + y2) // 2),
            (x2, (y1 + y2) // 2),
        ]

        for cx, cy in handles:
            cv2.circle(img_display, (cx, cy), 8, (255, 255, 0), -1)

def get_handle(x, y, box):
    x1, y1, x2, y2 = box

    # 🔥 FIX: stable + larger click area
    tol = max(12, int(25 / zoom_scale))  

    handles = {
        "top":    ((x1 + x2) // 2, y1),
        "bottom": ((x1 + x2) // 2, y2),
        "left":   (x1, (y1 + y2) // 2),
        "right":  (x2, (y1 + y2) // 2),
    }

    for name, (hx, hy) in handles.items():
        if abs(x - hx) <= tol and abs(y - hy) <= tol:
            return name

    return None
    
# =========================
# MOUSE EVENTS (CUT + EDIT)
# =========================
def on_mouse(event):
    global start_x, start_y, dragging
    global move_mode, resize_mode, resize_corner
    global selected_box_index, selected_mark_index
    global mark_mode
    global selected_mark_source   # 🔥 IMPORTANT FIX

    if img_cv is None:
        return

    # =========================
    # SELECT MODE (STRICT FIXED)
    # =========================
    if select_mode:
    
        x, y = screen_to_image(event.x, event.y)
    
        if event.type == tk.EventType.ButtonPress:
    
            selected_mark_index = None
            selected_box_index = None
            selected_mark_source = None
    
            # ================= BOX CHECK FIRST =================
            idx = find_box(x, y)
            if idx is not None:
                selected_box_index = idx
            
                box, *_ = history_stack[idx]
            
                # =========================
                # SHOW SIZE IN FRAME_BOX2
                # =========================
                x1, y1, x2, y2 = box
            
                w_px = abs(x2 - x1)
                h_px = abs(y2 - y1)
            
                w_cm = round_cm_custom(w_px * CM_PER_PIXEL)
                h_cm = round_cm_custom(h_px * CM_PER_PIXEL)
            
                h_entry2.delete(0, tk.END)
                h_entry2.insert(0, f"{h_cm:.1f}")
            
                w_entry2.delete(0, tk.END)
                w_entry2.insert(0, f"{w_cm:.1f}")
            
                update_preview_box(box)
            
                redraw()
                return # 🔥 STOP HERE (box has priority)
    
            # ================= MARK SELECTION (ONLY IF NO BOX) =================
    
            # GLOBAL MARKS
            for i, (mx, my) in enumerate(global_marks):
                if abs(x - mx) < 10 and abs(y - my) < 10:
                    selected_mark_index = i
                    selected_mark_source = "global"
                    redraw()
                    return
    
            # PAGE MARKS (BLUE +)
            for i, (mx, my) in enumerate(page_marks.get(page_index, [])):
                if abs(x - mx) < 10 and abs(y - my) < 10:
                    selected_mark_index = i
                    selected_mark_source = "page"
                    redraw()
                    return
    
            # ================= NOTHING FOUND =================
            update_preview_box(None)
            redraw()
            return

        return

    # =========================
    # BUTTON PRESS
    # =========================
    if event.type == tk.EventType.ButtonPress:

        x, y = screen_to_image(event.x, event.y)

        selected_mark_index = None
        selected_box_index = None
        selected_mark_source = None   # 🔥 RESET FIX

        # ================= MARK SELECTION (GLOBAL) =================
        if selected_box_index is None:
            for i, (mx, my) in enumerate(global_marks):
                if abs(x - mx) < 10 and abs(y - my) < 10:
                    selected_mark_index = i
                    selected_mark_source = "global"
                    redraw()
                    return

        # ================= BOX SELECTION =================
        idx = find_box(x, y)

        if idx is not None:
            selected_box_index = idx
        
            box, obj_ids, selected_marks, *rest = history_stack[idx]
        
            added_marks = rest[0] if rest else []
        
            # 🔥 ONLY select marks linked to THIS box
            selected_mark_index = None
            selected_mark_source = None
        
            # store locally for this selection
            current_box_marks = added_marks + selected_marks
        
            update_preview_box(box)
        
        else:
            selected_box_index = None
            update_preview_box(None)

        if edit_enabled and idx is not None:
            selected_box_index = idx

            box = history_stack[idx][0]

            edge = get_resize_direction(x, y, box)
            if edge:
                resize_mode = True
                resize_corner = edge
                dragging = False
                redraw()
                return

            redraw()
            return

        # ================= MARK MODE =================
        if mark_mode:
            global_marks.append((x, y))
            history_stack.append(("mark", (x, y)))
            redraw()
            return

        # ================= MARK SELECTION (PAGE / BLUE +) =================
        for i, (mx, my) in enumerate(page_marks.get(page_index, [])):
            if abs(x - mx) < 10 and abs(y - my) < 10:
                selected_mark_index = i
                selected_mark_source = "page"
                redraw()
                return

        # ================= CUT MODE START =================
        if mode == "cut" and not edit_enabled:
            move_mode = False
            resize_mode = False

            start_x, start_y = event.x, event.y
            dragging = True

    # =========================
    # MOTION
    # =========================
    elif event.type == tk.EventType.Motion:

        # ================= RESIZE =================
        if edit_enabled and resize_mode and selected_box_index is not None:

            x, y = screen_to_image(event.x, event.y)

            box, obj_ids, selected_marks, *rest = history_stack[selected_box_index]
            x1, y1, x2, y2 = box

            if resize_corner == "tl":
                x1, y1 = x, y
            elif resize_corner == "tr":
                x2, y1 = x, y
            elif resize_corner == "bl":
                x1, y2 = x, y
            elif resize_corner == "br":
                x2, y2 = x, y
            elif resize_corner == "left":
                x1 = x
            elif resize_corner == "right":
                x2 = x
            elif resize_corner == "top":
                y1 = y
            elif resize_corner == "bottom":
                y2 = y

            x1, x2 = min(x1, x2), max(x1, x2)
            y1, y2 = min(y1, y2), max(y1, y2)

            history_stack[selected_box_index] = ((x1, y1, x2, y2), obj_ids, selected_marks)
            update_preview_box((x1, y1, x2, y2))
            redraw()
            return

        # ================= DRAW PREVIEW =================
        if dragging and not edit_enabled and not select_mode:

            temp_img = img_cv.copy()

            for item in history_stack:
                if item[0] == "mark":
                    continue

                box, obj_ids, selected_marks, *rest = item
                x1, y1, x2, y2 = box

                cv2.rectangle(temp_img, (x1, y1), (x2, y2), (0, 0, 255), 2)

            # GLOBAL MARKS
            px = int((1 / CM_PER_PIXEL) * zoom_scale)

            for x, y in global_marks:
                cv2.line(temp_img, (x - px // 2, y), (x + px // 2, y), (0, 0, 255), 2)
                cv2.line(temp_img, (x, y - px // 2), (x, y + px // 2), (0, 0, 255), 2)

            # PAGE MARKS (BLUE +)
            for x, y in page_marks.get(page_index, []):
                cv2.line(temp_img, (x - px // 2, y), (x + px // 2, y), (255, 0, 0), 2)
                cv2.line(temp_img, (x, y - px // 2), (x, y + px // 2), (255, 0, 0), 2)

            x1, y1 = screen_to_image(start_x, start_y)
            x2, y2 = screen_to_image(event.x, event.y)

            cv2.rectangle(temp_img, (x1, y1), (x2, y2), (0, 255, 0), 2)

            img_rgb = cv2.cvtColor(temp_img, cv2.COLOR_BGR2RGB)
            img_pil = Image.fromarray(img_rgb)

            canvas_w = canvas.winfo_width()
            canvas_h = canvas.winfo_height()

            img_h, img_w = img_cv.shape[:2]
            fit_scale = min(canvas_w / img_w, canvas_h / img_h)
            final_scale = fit_scale * zoom_scale

            new_w = int(img_w * final_scale)
            new_h = int(img_h * final_scale)

            img_pil = img_pil.resize((new_w, new_h), Image.Resampling.LANCZOS)

            imgtk = ImageTk.PhotoImage(img_pil)

            canvas.image = imgtk

            x_center = canvas_w // 2
            y_center = canvas_h // 2

            canvas.delete("all")
            canvas.create_image(x_center, y_center, anchor="center", image=imgtk)

            canvas.config(scrollregion=(
                x_center - new_w // 2,
                y_center - new_h // 2,
                x_center + new_w // 2,
                y_center + new_h // 2
            ))

    # =========================
    # RELEASE
    # =========================
    elif event.type == tk.EventType.ButtonRelease:

        if mark_mode:
            return

        move_mode = False
        resize_mode = False
        resize_corner = None
        selected_box_index = None
        dragging = False

        x1, y1 = screen_to_image(start_x, start_y)
        x2, y2 = screen_to_image(event.x, event.y)

        if mode == "cut" and not edit_enabled and not select_mode:
            success = process_area(x1, y1, x2, y2)
            if not success:
                redraw()

customer_map = {}

def load_customer_map(file_path=None):
    global customer_map

    if not file_path:
        print("⚠ No customer file provided")
        return

    if not os.path.exists(file_path):
        print("⚠ File not found:", file_path)
        return

    wb = load_workbook(file_path)
    ws = wb.active

    customer_map.clear()

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
            
        key = str(row[0]).strip().replace(".0", "")
        value = str(row[1]).strip()
        
        customer_map[key] = value

    print("✔ Loaded:", file_path)

def get_file_number():
    if not current_file_path:
        return ""

    name = os.path.basename(current_file_path)
    parts = name.split("_")

    best_match = ""

    for i in range(1, len(parts) + 1):
        candidate = "_".join(parts[:i]).strip().replace(".0", "")

        if candidate in customer_map:
            best_match = candidate

    return best_match

# =========================
# SAVE (UNCHANGED)
# =========================
def smart_round_cm(value):
    return round_cm_custom(value)


def save_all():
    global saved_flag
    wb = Workbook()
    ws = wb.active

    # 🔥 BLOCK SAVE IF CURRENT PAGE NOT DONE
    if has_unsaved_changes():
        set_status("⚠ Click DONE before saving", True)
        print("⚠ Please click DONE before saving")
        return
        
    base_dir = os.path.dirname(current_file_path)

    raw_name = os.path.splitext(os.path.basename(current_file_path))[0]
    base_name = raw_name.rsplit("_", 1)[0]

    excel_file = os.path.join(base_dir,"Patches", f"{base_name}.xlsx")
    output_folder = os.path.join(base_dir, "Patches")
    os.makedirs(output_folder, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.append(["Date","Customer Name","Job Description","Plate Type", "Color Type","No.of Color",
               "L(CM)", "W(CM)","Colors", "Qty.(cm2)","Editor"])
    pdf_pages = []

    for p_index, state in page_data.items():

        history = state["history"]

        if p_index == 0:
            page_name = "ALL_COLORS"
        else:
            page_name = (
                page_names[p_index - 1]
                if (p_index - 1) < len(page_names)
                else f"Page_{p_index}"
            )

        page = pdf_doc.load_page(p_index)
        pix = page.get_pixmap(dpi=300)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

        if pix.n == 4:
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)

        img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
        img = cv2.resize(img, (0, 0), fx=SCALE, fy=SCALE)

        # =========================
        # ADD 20 CM BORDER ONLY FOR OUTPUT
        # =========================
        pad_px = int(10 / CM_PER_PIXEL)
        
        img_out = cv2.copyMakeBorder(
            img,
            pad_px, pad_px, pad_px, pad_px,
            borderType=cv2.BORDER_CONSTANT,
            value=(255, 255, 255)  # white border
        )

        for item in history:
        
            # 🔥 skip mark entries
            if item[0] == "mark":
                continue
        
            box = item[0]
            x1, y1, x2, y2 = box
        
            w_px = x2 - x1
            h_px = y2 - y1
        
            w_cm = smart_round_cm(w_px * CM_PER_PIXEL)
            h_cm = smart_round_cm(h_px * CM_PER_PIXEL)
            area_cm2 = w_cm * h_cm

            file_no = get_file_number()
            customer_name = customer_map.get(file_no, "")  # lookup

            ws.append([
                datetime.now().strftime("%d-%m-%Y"),            # A
                customer_name,                                  # B (Customer Name)
                base_name,                                      # C (Job Description)
                plate_var.get(),                                # D (Plate Type)
                page_name,                                      # E (Color Type / Page Name)
                "",                                             # F (No.of Color)
                h_cm,                                           # G (L(CM))
                w_cm,                                           # H (W(CM))
                1,                                              # I (Colors)
                area_cm2,                                       # J (Qty.(cm2))
                editor_var.get()                                # K (Editor)
            ])

            cv2.rectangle(img_out, (x1, y1), (x2, y2), (0, 0, 255), 2)
            # =========================
            # SIZE LABEL
            # =========================
            label = f"{h_cm:.1f} x {w_cm:.1f} cm"
            
            text_x = x1
            text_y = y1
            
            # keep text inside image
            if text_y < 0:
                text_y = y1 + pad_px + 0
            
            # background box for readability
            (text_w, text_h), _ = cv2.getTextSize(
                label,
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                2
            )
            
            cv2.rectangle(
                img_out,
                (text_x - 3, text_y - text_h - 5),
                (text_x + text_w + 3, text_y + 3),
                (255, 255, 255),
                -1
            )
            
            cv2.putText(
                img_out,
                label,
                (text_x, text_y),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (0, 0, 255),
                2,
                cv2.LINE_AA
            )
            
            # =========================
            # DRAW GLOBAL MARKS (ALL PAGES)
            # =========================
            for mx, my in global_marks:
                if x1 <= mx <= x2 and y1 <= my <= y2:
            
                    px = 6
            
                    cv2.line(img_out,(mx - px, my),(mx + px, my),(0, 0, 255), 2)         
                    cv2.line(img_out,(mx, my - px),(mx, my + px),(0, 0, 255), 2)
                        
            # =========================
            # DRAW MOUNTING MARKS (ONLY THIS PAGE)
            # =========================
            for mx, my in page_marks.get(p_index, []):
                if x1 <= mx <= x2 and y1 <= my <= y2:
            
                    px = 6
            
                    cv2.line(img_out,(mx - px, my),(mx + px, my),(255, 0, 0), 2)            
                    cv2.line(img_out,(mx, my - px),(mx, my + px),(255, 0, 0), 2)
        # =========================
        # PAGE COLOR NAME HEADER
        # =========================
        header_text = page_name

        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 3
        thickness = 6

        (text_w, text_h), _ = cv2.getTextSize(
            header_text,
            font,
            font_scale,
            thickness
        )

        img_h, img_w = img_out.shape[:2]

        text_x = (img_w - text_w) // 2
        text_y = 150

        # white background for clarity
        cv2.rectangle(
            img_out,
            (text_x - 15, text_y - text_h - 10),
            (text_x + text_w + 15, text_y + 10),
            (255, 255, 255),
            -1
        )

        # BLACK TEXT
        cv2.putText(
            img_out,
            header_text,
            (text_x, text_y),
            font,
            font_scale,
            (0, 0, 0),   # black
            thickness,
            cv2.LINE_AA
        )

        # convert OpenCV → PIL
        img_rgb = cv2.cvtColor(img_out, cv2.COLOR_BGR2RGB)
        
        pil_img = Image.fromarray(img_rgb).convert("RGB")
        
        pdf_pages.append(pil_img)
    # =========================
    # CENTER ALIGN ALL CELLS
    # =========================
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
    
    # =========================
    # AUTO COLUMN WIDTH
    # =========================
    from openpyxl.utils import get_column_letter
    
    for column_cells in ws.columns:
    
        max_length = 0
        column = column_cells[0].column
    
        for cell in column_cells:
            try:
                cell_value = str(cell.value)
    
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
    
            except:
                pass
    
        adjusted_width = max_length + 5
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    # =========================
    # SAVE COMBINED PDF
    # =========================
    if pdf_pages:
    
        pdf_output = os.path.join(
            output_folder,
            f"{base_name}.pdf"
        )
    
        pdf_pages[0].save(
            pdf_output,
            save_all=True,
            append_images=pdf_pages[1:]
        )
    
        print("✅ PDF SAVED:", pdf_output)    
    wb.save(excel_file)

    print("✅ ALL PAGES SAVED SUCCESSFULLY")

    set_status("Saved All ✅")
    saved_flag = True
    

def full_reset():
    global img_cv, img_display, pdf_doc
    global history_stack, redo_stack, selected_ids
    global global_marks, page_marks, page_data
    global selected_mark_ids
    global page_index, total_pages
    global current_file_path, saved_flag

    img_cv = None
    img_display = None
    pdf_doc = None

    history_stack.clear()
    redo_stack.clear()
    selected_ids.clear()

    global_marks.clear()
    page_marks.clear()
    page_data.clear()
    selected_mark_ids.clear()

    page_index = 0
    total_pages = 0
    current_file_path = None
    saved_flag = True

    canvas.delete("all")
    page_label_var.set("Color Name")

    print("🧹 FULL RESET DONE")

CONFIG_FILE = "customer_link_config.txt"

def load_customer_Excel_file():
    global customer_map
    global excel_path   # ADD THIS

    file_path = filedialog.askopenfilename(
        title="Select Customer Mapping Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if not file_path:
        return

    # save excel path
    excel_path = file_path   # ADD THIS

    # load immediately
    load_customer_map(file_path)

    # 🔥 ADD THESE
    load_plate_list()
    load_editor_list()

    update_plate_menu()
    update_editor_menu()

    # save path permanently
    with open(CONFIG_FILE, "w") as f:
        f.write(file_path)

    messagebox.showinfo("Success", "Customer file linked ✔")
    
# =========================
# UI
# =========================
root = tk.Tk()
root.title("Nurak Patcher")
root.state("zoomed")
mode_var = tk.StringVar(value="No Mark")
file_name_var = tk.StringVar(value="No File Loaded")
excel_path = None

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

root.iconbitmap(resource_path("AppLogo.ico"))

def update_preview_box(box):
    if box is None:
        sidebar_top_label.config(text="Patch Size")
        return

    x1, y1, x2, y2 = box
    
    h_px = abs(y2 - y1)
    w_px = abs(x2 - x1)
    
    h_cm = h_px * CM_PER_PIXEL
    w_cm = w_px * CM_PER_PIXEL
    # ✅ APPLY YOUR NEW RULE HERE ONLY    
    h_cm = round_cm_custom(h_cm)
    w_cm = round_cm_custom(w_cm)

    sidebar_top_label.config(
        text=f"H: {h_cm:.1f} cm\nW: {w_cm:.1f} cm"
    )

top_bar = tk.Frame(root, bg="#007B85", height=45)
top_bar.pack(side="top", fill="x")
top_bar.pack_propagate(False)

# =========================
# RESIZE SELECTED BOX FROM FRAME_BOX2
# =========================
def apply_box2_resize(event=None):

    global history_stack, selected_box_index

    # no selected box
    if selected_box_index is None:
        return

    if selected_box_index >= len(history_stack):
        return

    item = history_stack[selected_box_index]

    # skip mark entries
    if item[0] == "mark":
        return

    box = item[0]
    obj_ids = item[1] if len(item) > 1 else set()
    selected_marks = item[2] if len(item) > 2 else []
    added_marks = item[3] if len(item) > 3 else []

    x1, y1, x2, y2 = box

    # current center
    cx = (x1 + x2) // 2
    cy = (y1 + y2) // 2

    # current size fallback
    current_w_px = abs(x2 - x1)
    current_h_px = abs(y2 - y1)

    # =========================
    # GET NEW VALUES
    # =========================
    try:
        w_cm = float(w_entry2.get())
        new_w_px = int(w_cm / CM_PER_PIXEL)
    except:
        new_w_px = current_w_px

    try:
        h_cm = float(h_entry2.get())
        new_h_px = int(h_cm / CM_PER_PIXEL)
    except:
        new_h_px = current_h_px

    # =========================
    # BUILD NEW BOX
    # =========================
    new_x1 = int(cx - new_w_px / 2)
    new_x2 = int(cx + new_w_px / 2)

    new_y1 = int(cy - new_h_px / 2)
    new_y2 = int(cy + new_h_px / 2)

    # preserve your tuple structure
    history_stack[selected_box_index] = (
        (new_x1, new_y1, new_x2, new_y2),
        obj_ids,
        selected_marks,
        added_marks
    )

    update_preview_box((new_x1, new_y1, new_x2, new_y2))
    redraw()

# =========================
# ENABLE / DISABLE FRAME_BOX2
# =========================
def set_framebox2_state(enabled=True):

    state = "normal" if enabled else "disabled"

    for widget in frame_box2.winfo_children():
        try:
            widget.config(state=state)
        except:
            pass
            

tk.Button(top_bar, text="Open", command=open_file).pack(side="left", padx=(10,2), pady=5)
tk.Button(top_bar, text="Close", command=close_file).pack(side="left", padx=(2,2), pady=5)
tk.Button(top_bar, text="Auto cut", command=auto_cut).pack(side="left", padx=(2,2), pady=5)
def set_cut_mode():
    reset_all_modes()
    print("Cut mode ON")

tk.Button(top_bar, text="Cut", command=set_cut_mode).pack(side="left", padx=(2,2), pady=5)

def reset_all_modes():
    global select_mode, edit_enabled, mark_mode, mode

    select_mode = False
    edit_enabled = False
    mark_mode = False
    mode = "cut"   # 🔥 ensures cut is default fallback state

    # reset UI buttons
    select_btn.config(bg=default_select_bg, fg="black")
    edit_btn.config(bg=default_bg, fg="black")
    mark_btn.config(bg=default_mark_bg, fg="black")

    # 🔥 CLEAR FRAME_BOX2 WHEN SELECT MODE OFF
    h_entry2.delete(0, tk.END)
    w_entry2.delete(0, tk.END)

    update_preview_box(None)  

    set_framebox2_state(False)
    
def toggle_select_mode():
    global select_mode

    reset_all_modes()   # 🔥 turn everything OFF first

    select_mode = True  # activate only this

    select_btn.config(bg="black", fg="white")
    set_framebox2_state(True)
    print("Select mode ON")

select_btn = tk.Button(top_bar, text="⬩➤", command=toggle_select_mode)
select_btn.pack(side="left", padx=(2,2), pady=5)

default_select_bg = select_btn.cget("bg")

edit_enabled = False

def toggle_edit():
    global edit_enabled

    reset_all_modes()   # 🔥

    edit_enabled = True

    edit_btn.config(bg="green", fg="white")
    print("Edit mode ON")
    redraw()


edit_btn = tk.Button(top_bar, text="Edit", command=toggle_edit)
edit_btn.pack(side="left", padx=(2,2), pady=5)

default_bg = edit_btn.cget("bg")

def set_mark_mode():
    global mark_mode, mode

    # 🚫 BLOCK MARK MODE in Mounting Mark mode
    if mode_var.get() == "Mounting Mark":
        print("Mark mode blocked in Mounting Mark mode")
        return

    reset_all_modes()

    mark_mode = True
    mode = "mark"   # 🔥 ADD THIS

    mark_btn.config(bg="blue", fg="white")
    print("Mark mode ON")


mark_btn = tk.Button(top_bar, text="➕Mark", command=set_mark_mode)
mark_btn.pack(side="left", padx=(2,2), pady=5)
default_mark_bg = mark_btn.cget("bg")

tk.Button(top_bar, text="Undo", command=undo).pack(side="left", padx=(2,2), pady=5)
tk.Button(top_bar, text="Redo", command=redo).pack(side="left", padx=(2,10), pady=5)

frame_box = tk.Frame(top_bar)
frame_box.pack(side="left", padx=(0,10), pady=8)

tk.Label(frame_box, text="Margin/CM",font=("Arial",12,"bold")).pack(side="left")
tk.Label(frame_box, text="H:").pack(side="left")

h_entry = tk.Entry(frame_box, width=3, bd=1, highlightthickness=1)
h_entry.pack(side="left", ipady=2, padx=(2,3))
h_entry.insert(0, "1")

tk.Label(frame_box, text="W:").pack(side="left")

w_entry = tk.Entry(frame_box, width=3, bd=1, highlightthickness=1)
w_entry.pack(side="left", ipady=2, padx=(2,3))
w_entry.insert(0, "1")

# =========================
# CARET HANDLING (FIXED)
# =========================
def hide_caret(event=None):
    h_entry.selection_clear()
    w_entry.selection_clear()
    root.focus()   # remove cursor focus from entry boxes


# Bind Enter key
h_entry.bind("<Return>", hide_caret)
w_entry.bind("<Return>", hide_caret)

def done_clicked():
    global done_click_time

    now = time.time()

    # double click = within 0.5 sec
    if now - done_click_time < 0.5:
        save_and_next()
    else:
        save_current_only()

    done_click_time = now

def save_current_only():
    global page_index, edit_enabled, mark_mode, saved_flag

    saved_flag = False

    # turn off modes
    edit_enabled = False
    mark_mode = False

    # reset UI buttons
    edit_btn.config(bg=default_bg, fg="black")
    mark_btn.config(bg=default_mark_bg, fg="black")

    # save current page state
    page_data[page_index] = {
        "history": history_stack.copy(),
        "redo": redo_stack.copy(),
        "selected": selected_ids.copy(),
        "marks": global_marks.copy()
    }

    print(f"Page {page_index} saved ✔")

def save_and_next():
    global page_index

    save_current_only()

    if pdf_doc and page_index < total_pages - 1:
        save_current_page_state()
        page_index += 1
        load_pdf_page(page_index)

tk.Button(top_bar, text="Done", command=done_clicked, 
          bg="green", fg="white").pack(side="left", padx=(2,2), pady=5)

def reset_current_page():
    global history_stack, redo_stack, selected_ids
    global global_marks, page_marks, selected_mark_ids

    # Clear all boxes
    history_stack.clear()
    redo_stack.clear()
    selected_ids.clear()

    # Clear global marks
    selected_mark_ids.clear()

    # Clear page marks, but keep only red plus marks
    if page_index in page_marks:
        # red plus marks are tuples with length 3 and third element is "red_plus"
        page_marks[page_index] = [
            mark for mark in page_marks[page_index] if len(mark) == 3 and mark[2] == "red_plus"
        ]

    # Clear saved page state, but preserve red plus marks
    if page_index in page_data:
        page_data[page_index] = {
            "history": [],
            "redo": [],
            "selected": set(),
            "marks": [
                mark for mark in page_data[page_index].get("marks", [])
                if len(mark) == 3 and mark[2] == "red_plus"
            ]
        }

    # Reset any preview box
    update_preview_box(None)
    redraw()

    print(f"Page {page_index} RESET ✔ (red plus marks preserved)")

tk.Button(top_bar, text="➕", command=zoom_out).pack(side="right", padx=2)
tk.Button(top_bar, text="➖", command=zoom_in).pack(side="right", padx=2)
tk.Button(top_bar, text="🔍⟲", command=reset_zoom).pack(side="right", padx=2)
tk.Button(top_bar, text="RESET", command=reset_current_page).pack(side="right", padx=2)
tk.Button(top_bar, text="LINK", command=load_customer_Excel_file).pack(side="right", padx=2)
tk.Button(top_bar, text="➡", command=next_page).pack(side="right", padx=2)
tk.Button(top_bar, text="⬅", command=prev_page).pack(side="right", padx=2)

def save_and_close_artwork():

    # =========================
    # NO FILE OPEN
    # =========================
    if not current_file_path:
        messagebox.showwarning(
            "No File Open",
            "Please open a file first."
        )
        return

    # =========================
    # VALIDATION
    # =========================
    plate_value = plate_var.get().strip()
    editor_value = editor_var.get().strip()
    
    if not plate_value:
        messagebox.showwarning(
            "Missing Plate Type",
            "Please select Plate Type before saving."
        )
        plate_menu.focus_set()
        return
    
    if not editor_value:
        messagebox.showwarning(
            "Missing Editor",
            "Please select an Editor before saving."
        )
        return

    # =========================
    # SAVE + CLOSE
    # =========================
    save_all()
    close_file()
    
    plate_var.set("")
    editor_var.set("")
    
    # hide caret / remove focus
    root.focus_set()

tk.Button(top_bar,text="Save&Close", command=save_and_close_artwork, bg="green",fg="white",
          activebackground="green", activeforeground="white").pack(side="right", padx=2)

page_label_var = tk.StringVar(value="Color Name")

center_frame = tk.Frame(top_bar, bg="#007B85")
center_frame.pack(side="left", expand=True)

file_name_row = tk.Frame(root, bg="#007B85", height=15)
file_name_row.pack(side="top", fill="x")
file_name_row.pack_propagate(False)
base_name_var = tk.StringVar(value="")

file_name_label = tk.Label(
    file_name_row,
    textvariable=base_name_var,
    bg="#007B85",
    fg="white",
    font=("Arial", 10, "bold")
)
file_name_label.pack(side="left", padx=4, pady=2)

def set_status(text, warning=False):
    page_label_var.set(text)

    if warning:
        page_label.config(fg="red")
    else:
        page_label.config(fg="blue")

# 🔥 bordered box frame
label_box = tk.Frame(center_frame, bg="#007B85", highlightbackground="#BDBDBD", highlightthickness=5)
label_box.pack(pady=3)

page_label = tk.Label(
    label_box,
    textvariable=page_label_var,
    bg="#e0e0de",
    fg="blue",
    font=("Arial", 18),
    padx=15,
    pady=5
)
page_label.pack()

def name_selected():
    print("Name selected clicked")
    
def on_mode_change(value):
    reset_all_modes()
    mode_var.set(value)
    print("Mode selected:", value)

mode_menu = tk.OptionMenu(
    top_bar,
    mode_var,
    "No Mark",
    "Mounting Mark",
    command=on_mode_change
)

mode_menu.config(
    bg="white",
    fg="black",
    activebackground="lightgray"
)

mode_menu.pack(side="right", padx=2)

# =========================
# RIGHT SIDEBAR
# =========================
sidebar = tk.Frame(root, width=120, bg="#007B85")
sidebar.pack(side="right", fill="y")
sidebar.pack_propagate(False)

tk.Frame(sidebar, height=8, bg="#01494f").pack(side="top", fill="x", pady=(0,5))

# TOP PREVIEW (keep separate, DO NOT destroy)
sidebar_top = tk.Frame(sidebar, bg="white", height=80, highlightbackground="#BDBDBD", highlightthickness=5)
sidebar_top.pack(side="top", fill="x", padx=5, pady=5)
sidebar_top.pack_propagate(False)

sidebar_top_label = tk.Label(sidebar_top, text="Patch Size", bg="white", fg="gray")
sidebar_top_label.pack(expand=True)


frame_box2 = tk.Frame(sidebar)
frame_box2.pack(side="top", pady=(0,10))

tk.Label(frame_box2, text="H:").pack(side="left")

h_entry2 = tk.Entry(frame_box2, width=5, bd=1, highlightthickness=1)
h_entry2.pack(side="left", ipady=2, padx=(2,3))

tk.Label(frame_box2, text="W:").pack(side="left")

w_entry2 = tk.Entry(frame_box2, width=5, bd=1, highlightthickness=1)
w_entry2.pack(side="left", ipady=2, padx=(2,3))

set_framebox2_state(False)


# =========================
# APPLY RESIZE + HIDE CARET
# =========================

def hide_caret2(event=None):
    h_entry2.selection_clear()
    w_entry2.selection_clear()
    root.focus() 
    
def apply_resize_and_hide(event=None):

    apply_box2_resize()
    hide_caret2()

    return "break"


# Bind Enter key
h_entry2.bind("<Return>", apply_resize_and_hide)
w_entry2.bind("<Return>", apply_resize_and_hide)

tk.Frame(sidebar, height=8, bg="#01494f").pack(side="top", fill="x", pady=(0,10))


plate_list = []


def load_plate_list():
    global plate_list
    plate_list = []

    try:
        path = excel_path if excel_path else "Ndigitec_Customer_Name.xlsx"

        wb = load_workbook(path)
        ws = wb.active

        values = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and len(row) > 7 and row[7]:
                values.append(str(row[7]).strip())

        plate_list = list(dict.fromkeys([v for v in values if v]))

    except Exception as g:
        print("plate load error:", g)
        plate_list = []
load_plate_list()

frame_plate = tk.Frame(sidebar)
frame_plate.pack(side="top", pady=(0,10))

tk.Label(frame_plate, text="Plate:").pack(side="left")

# ensure list is not empty
plate_var = tk.StringVar()
plate_var.set("")   # keep empty default

if plate_list and len(plate_list) > 0:
    plate_menu = tk.OptionMenu(frame_plate, plate_var, *plate_list)
else:
    plate_menu = tk.OptionMenu(frame_plate, plate_var, "")

plate_menu.config(width=10)
plate_menu.pack(side="left", ipady=2, padx=(2,3))

tk.Frame(sidebar, height=8, bg="#01494f").pack(side="top", fill="x", pady=(0,10))



editor_list = []

def load_editor_list():
    global editor_list
    editor_list = []

    try:
        path = excel_path if excel_path else "Ndigitec_Customer_Name.xlsx"

        wb = load_workbook(path)
        ws = wb.active

        values = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and len(row) > 5 and row[5]:
                values.append(str(row[5]).strip())

        editor_list = list(dict.fromkeys([v for v in values if v]))

    except Exception as e:
        print("Editor load error:", e)
        editor_list = []

load_editor_list()

frame_editor = tk.Frame(sidebar)
frame_editor.pack(side="top", pady=(0,10))

tk.Label(frame_editor, text="Editor:").pack(side="left")

# ensure list is not empty
editor_var = tk.StringVar()
editor_var.set("")   # keep empty

if editor_list and len(editor_list) > 0:
    editor_menu = tk.OptionMenu(frame_editor, editor_var, *editor_list)
else:
    editor_menu = tk.OptionMenu(frame_editor, editor_var, "")

editor_menu.config(width=10)
editor_menu.pack(side="left", ipady=2, padx=(2,3))

tk.Frame(sidebar, height=8, bg="#01494f").pack(side="top", fill="x", pady=(0,10))

def update_plate_menu():
    menu = plate_menu["menu"]
    menu.delete(0, "end")

    for item in plate_list:
        menu.add_command(
            label=item,
            command=lambda value=item: plate_var.set(value)
        )

def update_editor_menu():
    menu = editor_menu["menu"]
    menu.delete(0, "end")

    for item in editor_list:
        menu.add_command(
            label=item,
            command=lambda value=item: editor_var.set(value)
        )


# 🔥 NEW: container ONLY for buttons
sidebar_btn_frame = tk.Frame(sidebar, bg="#007B85")
sidebar_btn_frame.pack(side="top", fill="both", expand=True)

def remove_caret(event):
    widget = event.widget

    # If click is inside any Entry → do nothing
    if isinstance(widget, tk.Entry):
        return

    # otherwise remove focus from entries
    root.focus_set()
        

canvas_border = tk.Frame(root, bg="#BDBDBD", padx=10, pady=10)
canvas_border.pack(side="right", fill="both", expand=True)

canvas = tk.Canvas(canvas_border, bg="#EBEBEB", highlightthickness=0)
canvas.pack(fill="both", expand=True)


# LEFT CLICK → selection
canvas_border.bind("<Button-1>", on_mouse)
canvas_border.bind("<B1-Motion>", on_mouse)
canvas_border.bind("<ButtonRelease-1>", on_mouse)

canvas.bind("<Button-1>", on_mouse)
canvas.bind("<B1-Motion>", on_mouse)
canvas.bind("<ButtonRelease-1>", on_mouse)

# MOUSE WHEEL → zoom
canvas.bind("<MouseWheel>", on_mouse_wheel)

# RIGHT CLICK → pan
canvas.bind("<ButtonPress-3>", start_pan)
canvas.bind("<B3-Motion>", do_pan)
canvas.bind("<ButtonRelease-3>", stop_pan)



def delete_selected(event=None):
    global selected_box_index, selected_mark_index, selected_mark_source
    global history_stack, selected_ids, global_marks, page_marks

    deleted = False

    # =========================
    # DELETE BOX
    # =========================
    if selected_box_index is not None and selected_box_index < len(history_stack):
    
        item = history_stack.pop(selected_box_index)
        
        box = item[0]
        obj_ids = item[1] if len(item) > 1 else set()
        selected_marks = item[2] if len(item) > 2 else []
        added_marks = item[3] if len(item) > 3 else []
    
        selected_ids -= obj_ids

        for obj_id in obj_ids:
            if obj_id in selected_ids:
                selected_ids.remove(obj_id)
    
        # 🔥 REMOVE LINKED PAGE MARKS
        if page_index in page_marks:
            for mark in added_marks:
                if mark in page_marks[page_index]:
                    page_marks[page_index].remove(mark)
    
                # 🔥 ALSO REMOVE FROM GLOBAL (FIX)
                if mark in global_marks:
                    global_marks.remove(mark)
    
        # 🔥 FIX: release mark locks from deleted box (MUST be outside)
        for mx, my in selected_marks:
            selected_mark_ids.discard((mx, my))
    
        for mx, my in added_marks:
            selected_mark_ids.discard((mx, my))
    
        selected_box_index = None
        deleted = True

    # =========================
    # DELETE MARK
    # =========================
    elif selected_mark_index is not None and selected_mark_source is not None:

        if selected_mark_source == "global":
            if selected_mark_index < len(global_marks):
                global_marks.pop(selected_mark_index)

        elif selected_mark_source == "page":
            if selected_mark_index < len(page_marks.get(page_index, [])):
                page_marks[page_index].pop(selected_mark_index)

        selected_mark_index = None
        selected_mark_source = None
        deleted = True

    if deleted:
        redo_stack.clear()
        redraw()

root.bind("<Delete>", delete_selected)

def is_typing():
    widget = root.focus_get()
    return isinstance(widget, tk.Entry) or isinstance(widget, tk.Text)
    

def safe_tool(action):
    if is_typing():
        return  # BLOCK ALL SHORTCUTS WHILE TYPING

    action()
    return "break"

# SAFE BINDING (after canvas creation)
canvas.bind("<Button-1>", remove_caret, add="+")
canvas_border.bind("<Button-1>", remove_caret, add="+")
sidebar.bind("<Button-1>", remove_caret)
top_bar.bind("<Button-1>", remove_caret)

# Select (A)
root.bind("<a>", lambda e: safe_tool(toggle_select_mode))
root.bind("<A>", lambda e: safe_tool(toggle_select_mode))

# Cut (C)
root.bind("<c>", lambda e: safe_tool(set_cut_mode))
root.bind("<C>", lambda e: safe_tool(set_cut_mode))

# Edit (Alt + E)
root.bind("<Alt-e>", lambda e: safe_tool(toggle_edit))
root.bind("<Alt-E>", lambda e: safe_tool(toggle_edit))

# Mark (Shift + P)
root.bind("<Shift-p>", lambda e: safe_tool(set_mark_mode))
root.bind("<Shift-P>", lambda e: safe_tool(set_mark_mode))

# Auto Cut (Ctrl + A)
root.bind("<Control-a>", lambda e: safe_tool(auto_cut))
root.bind("<Control-A>", lambda e: safe_tool(auto_cut))

# Undo / Redo
root.bind("<Control-z>", lambda e: safe_tool(undo))
root.bind("<Control-Z>", lambda e: safe_tool(undo))

root.bind("<Control-Shift-z>", lambda e: safe_tool(redo))
root.bind("<Control-Shift-Z>", lambda e: safe_tool(redo))
root.bind("<Control-y>", lambda e: safe_tool(redo))

# Open
root.bind("<Control-o>", lambda e: safe_tool(open_file))
root.bind("<Control-O>", lambda e: safe_tool(open_file))

# Close
root.bind("<Control-w>", lambda e: safe_tool(close_file))
root.bind("<Control-W>", lambda e: safe_tool(close_file))

# 1. FIRST load saved config file path
def auto_load_customer_file():
    global excel_path

    if not os.path.exists(CONFIG_FILE):
        return

    with open(CONFIG_FILE, "r") as f:
        file_path = f.read().strip()

    if file_path and os.path.exists(file_path):
        excel_path = file_path
        load_customer_map(file_path)

        load_plate_list()
        load_editor_list()

        update_plate_menu()
        update_editor_menu()

auto_load_customer_file()

# 2. THEN initialize UI-dependent lists again (important safety refresh)
load_plate_list()
load_editor_list()

# 3. NOW build UI menus (already created, so just refresh them)
update_plate_menu()
update_editor_menu()

root.mainloop()